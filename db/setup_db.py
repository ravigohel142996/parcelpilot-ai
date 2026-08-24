import os
import sqlite3
import openpyxl

# Paths
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
XLSX_PATH = os.path.join(BASE_DIR, "data", "raw", "ParcelPilot_Assessment_Data.xlsx")
DB_PATH = os.path.join(BASE_DIR, "db", "parcelpilot.db")

def clean_cell_value(val):
    """Normalize excel values for sqlite ingestion."""
    if val is None:
        return None
    # Boolean conversions
    if isinstance(val, bool):
        return 1 if val else 0
    # Convert string representation of 'None' to actual None
    if isinstance(val, str) and val.strip() == "None":
        return None
    return val

def setup_database():
    print(f"Opening Excel workbook: {XLSX_PATH}...")
    if not os.path.exists(XLSX_PATH):
        raise FileNotFoundError(f"Source Excel workbook not found at {XLSX_PATH}")
        
    wb = openpyxl.load_workbook(XLSX_PATH)
    
    print(f"Connecting to SQLite database: {DB_PATH}...")
    # Ensure directory exists
    os.makedirs(os.path.dirname(DB_PATH), exist_ok=True)
    # Remove existing database if any to ensure fresh start
    if os.path.exists(DB_PATH):
        os.remove(DB_PATH)
        
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    
    # 1. Create metadata table
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS metadata (
            key TEXT PRIMARY KEY,
            value TEXT
        )
    """)
    # Populate metadata from README sheet
    readme_sheet = wb["README"]
    for row in readme_sheet.iter_rows(min_row=1, values_only=True):
        if row[0] is not None:
            cursor.execute(
                "INSERT OR REPLACE INTO metadata (key, value) VALUES (?, ?)",
                (str(row[0]).strip(), clean_cell_value(row[1]))
            )
            
    # 2. Create accounts table
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS accounts (
            account_id TEXT PRIMARY KEY,
            account_name TEXT,
            plan TEXT,
            status TEXT,
            csm TEXT,
            contract_file TEXT,
            premium_support INTEGER,
            notes TEXT
        )
    """)
    accounts_sheet = wb["accounts"]
    # Skip header row
    for row in accounts_sheet.iter_rows(min_row=2, values_only=True):
        if row[0] is not None:
            cursor.execute("""
                INSERT INTO accounts (account_id, account_name, plan, status, csm, contract_file, premium_support, notes)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """, [clean_cell_value(cell) for cell in row])
            
    # 3. Create orders table
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS orders (
            order_id TEXT PRIMARY KEY,
            account_id TEXT,
            carrier TEXT,
            status TEXT,
            booked_at TEXT,
            pickup_window_start TEXT,
            pickup_window_end TEXT,
            pickup_actual_at TEXT,
            shipment_fee_inr REAL,
            carrier_fault INTEGER,
            customer_fault INTEGER,
            cancellation_requested_at TEXT,
            notes TEXT,
            FOREIGN KEY (account_id) REFERENCES accounts (account_id)
        )
    """)
    orders_sheet = wb["orders"]
    for row in orders_sheet.iter_rows(min_row=2, values_only=True):
        if row[0] is not None:
            cursor.execute("""
                INSERT INTO orders (order_id, account_id, carrier, status, booked_at, 
                                    pickup_window_start, pickup_window_end, pickup_actual_at, 
                                    shipment_fee_inr, carrier_fault, customer_fault, 
                                    cancellation_requested_at, notes)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, [clean_cell_value(cell) for cell in row])
            
    # 4. Create tickets table
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS tickets (
            ticket_id TEXT PRIMARY KEY,
            account_id TEXT,
            created_at TEXT,
            status TEXT,
            subject TEXT,
            description TEXT,
            channel TEXT,
            assigned_to TEXT,
            last_customer_message_at TEXT,
            historical_resolution TEXT,
            FOREIGN KEY (account_id) REFERENCES accounts (account_id)
        )
    """)
    tickets_sheet = wb["tickets"]
    for row in tickets_sheet.iter_rows(min_row=2, values_only=True):
        if row[0] is not None:
            cursor.execute("""
                INSERT INTO tickets (ticket_id, account_id, created_at, status, subject, 
                                     description, channel, assigned_to, last_customer_message_at, 
                                     historical_resolution)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, [clean_cell_value(cell) for cell in row])
            
    # 5. Create escalations table
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS escalations (
            escalation_id INTEGER PRIMARY KEY AUTOINCREMENT,
            ticket_id TEXT,
            priority TEXT,
            reason TEXT,
            escalated_at TEXT,
            FOREIGN KEY (ticket_id) REFERENCES tickets (ticket_id)
        )
    """)

            
    conn.commit()
    conn.close()
    print("Database schema created and populated successfully!")

if __name__ == "__main__":
    setup_database()
